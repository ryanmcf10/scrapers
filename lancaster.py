"""
Download election summary results from the Lancaster County Elections website. Save the results as an Excel file.

All of the pages are consistently formatted, but are a jumble of un-classed, un-ID'd tables. This script attempts to sort
through them by traversing HTML elements. If the layout of the pages is changed in the future, this script will be broken.
"""
from datetime import date

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

"""
Set ROOT_URL to one of the links at the site below:

http://vr.co.lancaster.pa.us/ElectionReturns/Election_Returns.html

e.g. Setting ROOT_URL to 'http://vr.co.lancaster.pa.us/ElectionReturns/November_5,_2019_-_Municipal_Election/Categories.html' 
will download all of the results for the November 5, 2019 General Election.
"""
ROOT_URL = "http://vr.co.lancaster.pa.us/ElectionReturns/November_3,_2020_-_General_Election/Categories.html"
IGNORED_LINKS = ["RETURN", "QUESTIONS", "CATEGORIES"]


def main():
    print("Downloading Lancaster County election results...")
    
    results = parse(ROOT_URL)
    save(results)

    print("Done.")


def parse(url):
    """
    Load the page at the given URL. If it contains results, parse and return them. Otherwise, recursively click through
    the links on the page until you get to a page containing results.
    """
    print(url)

    results = []
    
    page = requests.get(url)
    soup = BeautifulSoup(page.text, 'html.parser')
    
    if has_results(soup):
        results += parse_results(soup)
    else:
        for a in soup.find_all('a'):
            if a.text.upper() not in IGNORED_LINKS:
                results += parse(a['href'])
                
    return results


def has_results(soup):
    """
    Check if the current page contains a Results Table.
    """
    # All of the results tables are enclosed by <br />'s above and below...
    if len(soup.find_all('br')) != 2:
        return False

    # ... But some pages that don't have any results still have 2 <br />'s
    # If the results table is on the page, it is 2 'siblings' after the first <br />
    return (soup.find('br').next_sibling.next_sibling is not None and
            soup.find('br').next_sibling.next_sibling.name == 'table')


def parse_results(soup):
    """
    Parse the elements of the Results Table into a 2-D array.
    """
    results = []

    summary_table = soup.find_all('br')[0].next_sibling.next_sibling
    results_table = soup.find_all('br')[-1].previous_sibling.previous_sibling

    office = summary_table.find_all('tr')[0].text.strip()

    vote_for = [x.upper() for x in summary_table.find_all('tr')[-1].find_all('td')[-1].text.strip("()").split()]

    if "ONE" in vote_for:
        vote_for = 1
    elif "TWO" in vote_for:
        vote_for = 2
    elif "THREE" in vote_for:
        vote_for = 3
    elif "FOUR" in vote_for:
        vote_for = 4
    elif "FIVE" in vote_for:
        vote_for = 5

    row = [office, vote_for]

    for tr in results_table.find_all('tr'):

        if len(tr.find_all('td')) == 2:
            name = tr.find_all('td')[0].text
            votes = tr.find_all('td')[-1].text

            if name.upper() == "BY PRECINCT":
                break

            else:
                results.append(row + [name, int(votes)])

    return results


def save(results):
    """
    Save all of the results as an Excel Workbook.
    """
    wb = Workbook()
    sheet = wb.active

    header = ['Contest', 'Vote For', 'Candidate', 'Votes']
    for col, val in enumerate(header, start=1):
        sheet.cell(1, col).value = val

    for row_number, row in enumerate(results, start=2):
        for col_number, value in enumerate(row, start=1):
            sheet.cell(row=row_number, column=col_number).value = value

    filename = f'lancaster_election_results_{str(date.today()).replace("-", "")}.xlsx'

    wb.save(filename)

    print(f"Saved {filename}")


if __name__ == '__main__':
    main()